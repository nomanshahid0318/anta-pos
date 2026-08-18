#!/usr/bin/env python3
"""
Migrate data exported from Google Sheets into the ANTA POS SQLite/PostgreSQL DB.

Supported inputs:
  1) Directory of CSV files named after sheets:
       Sales.csv, Products.csv, Stores.csv, Users.csv, Banks.csv,
       Returns.csv, Exchanges.csv, Claims.csv, Expenses.csv,
       Inventory.csv, Store_GRN.csv, Supplier_GRN.csv, HO_Warehouse.csv
  2) Single Excel workbook (.xlsx) with sheet tabs matching those names.
  3) Live Apps Script web-app URL (optional, read-only pull).

Usage:
  python scripts/migrate_from_sheets.py --csv-dir ./export
  python scripts/migrate_from_sheets.py --xlsx ./ANTA_POS.xlsx
  python scripts/migrate_from_sheets.py --apps-script-url URL --key ANTA2026

PINs in Users sheet are re-hashed. Existing barcodes/invoice IDs are upserted
(no duplicates). Run against a fresh or existing database safely.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

# Make backend importable
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.auth import hash_pin  # noqa: E402
from app.database import SessionLocal, init_db  # noqa: E402
from app.models import (  # noqa: E402
    Bank,
    Claim,
    Exchange,
    Expense,
    HOWarehouse,
    Inventory,
    Product,
    Return,
    Sale,
    Store,
    StoreGRN,
    SupplierGRN,
    User,
)


def _s(row: dict, *keys: str, default: str = "") -> str:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return str(row[k]).strip()
    # case-insensitive
    lower = {str(k).lower(): v for k, v in row.items()}
    for k in keys:
        if k.lower() in lower and lower[k.lower()] not in (None, ""):
            return str(lower[k.lower()]).strip()
    return default


def _f(row: dict, *keys: str, default: float = 0.0) -> float:
    v = _s(row, *keys, default="")
    try:
        return float(v) if v != "" else default
    except ValueError:
        return default


def _i(row: dict, *keys: str, default: int = 0) -> int:
    return int(_f(row, *keys, default=float(default)))


def load_csv_dir(d: Path) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for p in d.glob("*.csv"):
        with p.open(newline="", encoding="utf-8-sig") as f:
            out[p.stem] = list(csv.DictReader(f))
    return out


def load_xlsx(path: Path) -> dict[str, list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    out: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        data = []
        for r in rows[1:]:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            data.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
        out[name] = data
    return out


def pull_apps_script(url: str, key: str) -> dict[str, list[dict]]:
    import urllib.parse
    import urllib.request

    actions = [
        "products",
        "stores",
        "banks",
        "users",
        "sales",
        "inventory",
        "grns",
        "expenses",
        "warehouse",
    ]
    # map action -> sheet-like name
    rename = {
        "grns": "Store_GRN",
        "warehouse": "HO_Warehouse",
        "products": "Products",
        "stores": "Stores",
        "banks": "Banks",
        "users": "Users",
        "sales": "Sales",
        "inventory": "Inventory",
        "expenses": "Expenses",
    }
    out: dict[str, list[dict]] = {}
    base = url.split("?")[0]
    for action in actions:
        q = urllib.parse.urlencode({"action": action, "key": key})
        req = urllib.request.Request(f"{base}?{q}")
        with urllib.request.urlopen(req, timeout=60) as resp:  # noqa: S310
            payload = json.loads(resp.read().decode("utf-8"))
        if payload.get("ok") and isinstance(payload.get("data"), list):
            out[rename.get(action, action)] = payload["data"]
            print(f"  pulled {action}: {len(payload['data'])} rows")
        else:
            print(f"  skip {action}: {payload}")
    return out


def migrate(data: dict[str, list[dict]]) -> None:
    init_db()
    db = SessionLocal()
    stats = {k: 0 for k in [
        "stores", "users", "banks", "products", "sales", "returns",
        "exchanges", "claims", "expenses", "inventory", "store_grn",
        "supplier_grn", "warehouse",
    ]}
    try:
        for row in data.get("Stores", []):
            sid = _s(row, "StoreID", "store_id", "storeId")
            if not sid:
                continue
            existing = db.query(Store).filter(Store.store_id == sid).first()
            if existing:
                existing.name = _s(row, "Name", "name", default=existing.name)
                existing.city = _s(row, "City", "city", default=existing.city or "")
            else:
                db.add(
                    Store(
                        store_id=sid,
                        name=_s(row, "Name", "name", default=sid),
                        city=_s(row, "City", "city"),
                        address=_s(row, "Address", "address"),
                        manager=_s(row, "Manager", "manager"),
                        phone=_s(row, "Phone", "phone"),
                        active=_s(row, "Active", default="Y").upper() != "N",
                    )
                )
            stats["stores"] += 1

        for row in data.get("Users", []):
            uid = _s(row, "UserID", "user_id", "userId")
            pin = _s(row, "PIN", "pin", default="0000")
            if not uid:
                continue
            existing = db.query(User).filter(User.user_id == uid).first()
            if existing:
                existing.name = _s(row, "Name", "name", default=existing.name)
                existing.role = _s(row, "Role", "role", default=existing.role)
                existing.store_id = _s(row, "StoreID", "store_id", default=existing.store_id)
                existing.store_name = _s(row, "StoreName", "store_name", default=existing.store_name)
                if pin:
                    existing.pin_hash = hash_pin(pin)
            else:
                db.add(
                    User(
                        user_id=uid,
                        store_id=_s(row, "StoreID", "store_id"),
                        store_name=_s(row, "StoreName", "store_name"),
                        name=_s(row, "Name", "name", default=uid),
                        role=_s(row, "Role", "role", default="cashier"),
                        pin_hash=hash_pin(pin),
                        active=_s(row, "Active", default="Y").upper() != "N",
                    )
                )
            stats["users"] += 1

        for row in data.get("Banks", []):
            bid = _s(row, "BankID", "bank_id", "bankId") or f"B{stats['banks']+1}"
            existing = db.query(Bank).filter(Bank.bank_id == bid).first()
            name = _s(row, "Name", "name", default=bid)
            if existing:
                existing.name = name
                existing.device = _s(row, "Device", "device", default=existing.device or "")
            else:
                db.add(
                    Bank(
                        bank_id=bid,
                        name=name,
                        account_no=_s(row, "AccountNo", "account_no"),
                        device=_s(row, "Device", "device"),
                        active=_s(row, "Active", default="Y").upper() != "N",
                        icon="💵" if name.lower() == "cash" else "🏦",
                    )
                )
            stats["banks"] += 1

        for row in data.get("Products", []):
            bc = _s(row, "Barcode", "barcode")
            if not bc:
                continue
            existing = db.query(Product).filter(Product.barcode == bc).first()
            fields = dict(
                name=_s(row, "Name", "name", default=bc),
                brand=_s(row, "Brand", "brand", default="ANTA"),
                category=_s(row, "Category", "category"),
                size=_s(row, "Size", "size"),
                cost=_f(row, "Cost", "cost"),
                retail=_f(row, "Retail", "retail"),
                reorder=_i(row, "Reorder", "reorder", default=5),
                opening=_i(row, "Opening", "opening"),
                active=_s(row, "Active", default="Y").upper() != "N",
            )
            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
            else:
                db.add(Product(barcode=bc, **fields))
            stats["products"] += 1

        for row in data.get("Sales", []):
            inv = _s(row, "InvoiceID", "invoice_id", "id")
            store_id = _s(row, "StoreID", "store_id", "storeId")
            if not inv:
                continue
            if db.query(Sale).filter(Sale.invoice_id == inv, Sale.store_id == store_id).first():
                continue
            items_raw = row.get("Items_JSON") or row.get("items_json") or row.get("items") or "[]"
            if not isinstance(items_raw, str):
                items_raw = json.dumps(items_raw)
            db.add(
                Sale(
                    invoice_id=inv,
                    date=str(_s(row, "Date", "date"))[:10],
                    time=_s(row, "Time", "time"),
                    store=_s(row, "Store", "store"),
                    store_id=store_id,
                    customer=_s(row, "Customer", "customer", default="Walk-in"),
                    items_json=items_raw,
                    subtotal=_f(row, "Subtotal", "subtotal"),
                    discount=_f(row, "Discount", "discount"),
                    total=_f(row, "Total", "total"),
                    payment=_s(row, "Payment", "payment", default="Cash"),
                    pay_ref=_s(row, "PayRef", "pay_ref", "payRef"),
                    type=_s(row, "Type", "type", default="sale"),
                )
            )
            stats["sales"] += 1

        for row in data.get("Returns", []):
            ref = _s(row, "RefID", "ref_id", "ref")
            if not ref or db.query(Return).filter(Return.ref_id == ref).first():
                continue
            db.add(
                Return(
                    ref_id=ref,
                    date=str(_s(row, "Date", "date"))[:10],
                    time=_s(row, "Time", "time"),
                    store=_s(row, "Store", "store"),
                    store_id=_s(row, "StoreID", "store_id"),
                    orig_invoice=_s(row, "OrigInvoice", "orig_invoice"),
                    barcode=_s(row, "Barcode", "barcode"),
                    product_name=_s(row, "ProductName", "product_name"),
                    qty=_i(row, "Qty", "qty", default=1),
                    amount=_f(row, "Amount", "amount"),
                    method=_s(row, "Method", "method", default="Cash"),
                    reason=_s(row, "Reason", "reason"),
                )
            )
            stats["returns"] += 1

        for row in data.get("Exchanges", []):
            ref = _s(row, "RefID", "ref_id", "ref")
            if not ref or db.query(Exchange).filter(Exchange.ref_id == ref).first():
                continue
            db.add(
                Exchange(
                    ref_id=ref,
                    date=str(_s(row, "Date", "date"))[:10],
                    time=_s(row, "Time", "time"),
                    store=_s(row, "Store", "store"),
                    store_id=_s(row, "StoreID", "store_id"),
                    customer=_s(row, "Customer", "customer"),
                    old_barcode=_s(row, "OldBarcode", "old_barcode"),
                    old_name=_s(row, "OldName", "old_name"),
                    old_qty=_i(row, "OldQty", "old_qty", default=1),
                    new_barcode=_s(row, "NewBarcode", "new_barcode"),
                    new_name=_s(row, "NewName", "new_name"),
                    new_qty=_i(row, "NewQty", "new_qty", default=1),
                    diff=_f(row, "Diff", "diff"),
                    payment=_s(row, "Payment", "payment", default="Cash"),
                )
            )
            stats["exchanges"] += 1

        for row in data.get("Claims", []):
            ref = _s(row, "RefID", "ref_id", "ref")
            if not ref or db.query(Claim).filter(Claim.ref_id == ref).first():
                continue
            db.add(
                Claim(
                    ref_id=ref,
                    date=str(_s(row, "Date", "date"))[:10],
                    time=_s(row, "Time", "time"),
                    store=_s(row, "Store", "store"),
                    store_id=_s(row, "StoreID", "store_id"),
                    barcode=_s(row, "Barcode", "barcode"),
                    product_name=_s(row, "ProductName", "product_name"),
                    qty=_i(row, "Qty", "qty", default=1),
                    type=_s(row, "Type", "type", default="Damage"),
                    value=_f(row, "Value", "value"),
                    supplier=_s(row, "Supplier", "supplier"),
                    notes=_s(row, "Notes", "notes"),
                )
            )
            stats["claims"] += 1

        for row in data.get("Expenses", []):
            eid = _s(row, "ExpID", "exp_id", "id")
            if not eid or db.query(Expense).filter(Expense.exp_id == eid).first():
                continue
            db.add(
                Expense(
                    exp_id=eid,
                    date=str(_s(row, "Date", "date"))[:10],
                    store_id=_s(row, "StoreID", "store_id", default="HO"),
                    store=_s(row, "Store", "store", default="HO"),
                    category=_s(row, "Category", "category"),
                    sub_category=_s(row, "SubCategory", "sub_category"),
                    description=_s(row, "Description", "description"),
                    amount=_f(row, "Amount", "amount"),
                    pay_method=_s(row, "PayMethod", "pay_method", default="Cash"),
                    reference=_s(row, "Reference", "reference"),
                    notes=_s(row, "Notes", "notes"),
                )
            )
            stats["expenses"] += 1

        for row in data.get("Inventory", []):
            bc = _s(row, "Barcode", "barcode")
            store_id = _s(row, "StoreID", "store_id")
            store = _s(row, "Store", "store")
            if not bc or not store_id:
                continue
            existing = (
                db.query(Inventory)
                .filter(Inventory.barcode == bc, Inventory.store_id == store_id)
                .first()
            )
            vals = dict(
                name=_s(row, "Name", "name"),
                store=store,
                grn_in=_i(row, "GRN_In", "grn_in"),
                sales_out=_i(row, "Sales_Out", "sales_out"),
                returns_in=_i(row, "Returns_In", "returns_in"),
                exch_out=_i(row, "ExchOut", "exch_out"),
                exch_in=_i(row, "ExchIn", "exch_in"),
                claims=_i(row, "Claims", "claims"),
                on_hand=_i(row, "OnHand", "on_hand"),
            )
            if existing:
                for k, v in vals.items():
                    setattr(existing, k, v)
                existing.recalc()
            else:
                inv = Inventory(barcode=bc, store_id=store_id, **vals)
                inv.recalc()
                db.add(inv)
            stats["inventory"] += 1

        for row in data.get("Store_GRN", []):
            db.add(
                StoreGRN(
                    grn_id=_s(row, "GRNID", "grn_id"),
                    date=str(_s(row, "Date", "date"))[:10],
                    store_id=_s(row, "StoreID", "store_id"),
                    store_name=_s(row, "StoreName", "store_name"),
                    barcode=_s(row, "Barcode", "barcode"),
                    name=_s(row, "Name", "name"),
                    qty_issued=_i(row, "QtyIssued", "qty_issued"),
                    qty_received=_i(row, "QtyReceived", "qty_received"),
                    status=_s(row, "Status", "status", default="pending"),
                    notes=_s(row, "Notes", "notes"),
                )
            )
            stats["store_grn"] += 1

        for row in data.get("Supplier_GRN", []):
            db.add(
                SupplierGRN(
                    grn_id=_s(row, "GRNID", "grn_id"),
                    date=str(_s(row, "Date", "date"))[:10],
                    supplier=_s(row, "Supplier", "supplier"),
                    invoice_no=_s(row, "InvoiceNo", "invoice_no"),
                    barcode=_s(row, "Barcode", "barcode"),
                    name=_s(row, "Name", "name"),
                    qty=_i(row, "Qty", "qty"),
                    unit_cost=_f(row, "UnitCost", "unit_cost"),
                    total_cost=_f(row, "TotalCost", "total_cost"),
                    notes=_s(row, "Notes", "notes"),
                )
            )
            stats["supplier_grn"] += 1

        for row in data.get("HO_Warehouse", []):
            bc = _s(row, "Barcode", "barcode")
            if not bc:
                continue
            existing = db.query(HOWarehouse).filter(HOWarehouse.barcode == bc).first()
            if existing:
                existing.supplier_in = _i(row, "Supplier_In", "supplier_in")
                existing.store_out = _i(row, "Store_Out", "store_out")
                existing.recalc()
            else:
                wh = HOWarehouse(
                    barcode=bc,
                    name=_s(row, "Name", "name"),
                    supplier_in=_i(row, "Supplier_In", "supplier_in"),
                    store_out=_i(row, "Store_Out", "store_out"),
                )
                wh.recalc()
                db.add(wh)
            stats["warehouse"] += 1

        db.commit()
        print("Migration complete:")
        for k, v in stats.items():
            if v:
                print(f"  {k}: {v}")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main():
    ap = argparse.ArgumentParser(description="Migrate Google Sheets data into ANTA POS DB")
    ap.add_argument("--csv-dir", type=Path, help="Folder of CSV exports")
    ap.add_argument("--xlsx", type=Path, help="Excel workbook export")
    ap.add_argument("--apps-script-url", type=str, help="Deployed Apps Script web app URL")
    ap.add_argument("--key", type=str, default="ANTA2026", help="Apps Script secret key")
    args = ap.parse_args()

    data: dict[str, list[dict]] = {}
    if args.csv_dir:
        print(f"Loading CSVs from {args.csv_dir}")
        data = load_csv_dir(args.csv_dir)
    elif args.xlsx:
        print(f"Loading XLSX {args.xlsx}")
        data = load_xlsx(args.xlsx)
    elif args.apps_script_url:
        print(f"Pulling from Apps Script...")
        data = pull_apps_script(args.apps_script_url, args.key)
    else:
        ap.error("Provide --csv-dir, --xlsx, or --apps-script-url")

    if not data:
        print("No data found.")
        sys.exit(1)
    print("Sheets:", ", ".join(f"{k}({len(v)})" for k, v in data.items()))
    migrate(data)


if __name__ == "__main__":
    main()
